# CopyPaste - copying and pasting specified ranges in template
import openpyxl as xl

# TODO: Loop receivedFile through content of specified folder. Folder to be selected by user.
receivedFile = ('P:\\Documents Svetlana\\Excel training\\Marcos\\Regional templates\\' +\
                                    'G&A_planning_template_Africa1_FCST2_2020.xlsm')

sourceWB = xl.load_workbook(receivedFile)
try:
    sourceWS = sourceWB.get_sheet_by_name('Summary')
except KeyError:
    print('ERROR: Wrong tab name in sourse file.') # Insert name of the file which cause an error
# TODO: To scip file with error, in the end print out list of all scipped files.

# TODO: To be selected by user.
templateFile = ('P:\\Documents Svetlana\\Excel training\\Marcos\\Regional templates\\' +\
                                                'G&A_planning_template_FCST2_2020.xlsm')
templateWB = xl.load_workbook(templateFile, keep_vba=True)
templateWS = templateWB.get_sheet_by_name('Summary')

# Copy and pase defined columns from received file to template.
columnList = [2, 5, 8, 9, 11, 12, 14, 15, 17, 19, 21, 23, 24, 26, 27, 28, 29, 30,
                                            31, 32, 33, 34, 35, 36, 37, 38, 39, 40, 41]
for column in columnList:
    for row in range(6, sourceWS.max_row):
        c = sourceWS.cell(row=row, column=column)
        templateWS.cell(row=row, column=column).value = c.value
# Delete all unfilled rows.
for row in range (templateWS.max_row, 6, -1):
    if templateWS.cell(row=row, column=5).value == 'Please select':
        templateWS.delete_rows(row)
# TODO: Save in new folder.
templateWB.save('P:\\Documents Svetlana\\Excel training\\Marcos\\Regional templates\\To be uploaded\\' +\
                                    'G&A_planning_template_NEW_FCST2_2020.xlsm')