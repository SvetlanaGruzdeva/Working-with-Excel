# Consolidate - copy and paste data from multiple files in one in order to create consolidated file.

import os, logging, openpyxl as xl, win32com.client as win32
from pathlib import Path
from tkinter.filedialog import askopenfilename, askdirectory
from openpyxl.styles import Font
from openpyxl.formula.translate import Translator
from copy import copy

# userFolder = 'P:\\Documents Svetlana\\Excel training\\Marcos\\Consolidate\\'
# templateFile = 'P:\\Documents Svetlana\\Excel training\\Marcos\\Consolidate\\G&A_planning_template_FCST2_2020.xlsm'
# TODO: Take folder address from file address
userFolder = (os.path.abspath(askdirectory()) + '\\')
templateFile = askopenfilename()

logging.disable(logging.CRITICAL)
logging.basicConfig(filename=(str(userFolder) + 'logs.txt'), level=logging.DEBUG, format=' %(asctime)s - %(levelname)s - %(message)s')

lastRow = 6
templateWB = xl.load_workbook(templateFile, keep_vba=True)
templateWS = templateWB.get_sheet_by_name('Summary')

for file in os.listdir(userFolder):
    fileName = os.fsdecode(file)
    if (fileName.endswith(".xlsm") or fileName.endswith(".xlsx")) and not fileName == Path(templateFile).name:
        logging.info(fileName)
        receivedFile = (userFolder + fileName)
        sourceWB = xl.load_workbook(receivedFile)
        sourceWS = sourceWB.get_sheet_by_name('Summary')
        # Copy and paste defined columns from received file to template.
        columnList = [2, 5, 8, 9, 11, 12, 14, 15, 17, 19, 21, 23, 24, 26, 27, 29]
        
        logging.debug(lastRow)
        logging.debug('Source file last row: ' + str(sourceWS.max_row))
        for row in range(6, (sourceWS.max_row + 1)):
            logging.debug(row)
            for column in columnList:
                c = sourceWS.cell(row=row, column=column)
                templateWS.cell(row=lastRow, column=column).value = c.value
                templateWS.cell(row=lastRow, column=column).font = copy(c.font)
                templateWS.cell(row=lastRow, column=column).border = copy(c.border)
                templateWS.cell(row=lastRow, column=column).fill = copy(c.fill)
                templateWS.cell(row=lastRow, column=column).number_format = copy(c.number_format)
                templateWS.cell(row=lastRow, column=column).protection = copy(c.protection)
                templateWS.cell(row=lastRow, column=column).alignment = copy(c.alignment)

            lastRow += 1
    # TODO: It's better to save file as consolidated by script, not manually before running script.
    # region = templateWS.cell(row=6, column=2).value
    # consolFile = templateFile.replace('G&A', 'Consol_' + region + '_G&A')
    templateWB.save(templateFile)

# Insert formulas accross whole file
fColumnList = [1, 3, 4, 6, 7, 10, 13, 16, 18, 20, 22, 30, 31, 32, 33, 34, 35, 36,
                                                                    37, 38, 30, 40, 41] # Formula Column List
logging.debug('Consolidated File last row: ' + str(templateWS.max_row))
for row in range(7, (templateWS.max_row + 1)):
    for column in fColumnList:
        c = templateWS.cell(row=6, column=column)
        copyCell = c.column_letter + str(6)
        copyFormula = templateWS.cell(row=6, column=column).value
        pasteCell = templateWS.cell(row=row, column=column).column_letter + str(row)
        templateWS.cell(row=row, column=column).value = Translator(copyFormula, origin=copyCell).translate_formula(pasteCell)
        templateWS.cell(row=row, column=column).font = copy(c.font)
        templateWS.cell(row=row, column=column).border = copy(c.border)
        templateWS.cell(row=row, column=column).fill = copy(c.fill)
        templateWS.cell(row=row, column=column).number_format = copy(c.number_format)
        templateWS.cell(row=row, column=column).protection = copy(c.protection)
        templateWS.cell(row=row, column=column).alignment = copy(c.alignment)
templateWB.save(templateFile)

# TODO: Add check if there are still rows to detele.
# Delete all unfilled rows.
# for row in range (templateWS.max_row, 6, -1):
#     if templateWS.cell(row=row, column=5).value == 'Please select':
#         templateWS.delete_rows(row)
#     templateWB.save(templateFile)

logging.info('END OF SESSION')